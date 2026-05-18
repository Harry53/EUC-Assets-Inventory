from __future__ import annotations

import csv
import io
import json
import os
import shutil
import smtplib
import sqlite3
from datetime import date, datetime, timedelta
from functools import wraps
from pathlib import Path
from typing import Any

from flask import Flask, Response, flash, redirect, render_template, request, session, url_for
from werkzeug.security import check_password_hash, generate_password_hash

from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
DB_PATH = BASE_DIR / "inventory.db"
LOG_PATH = BASE_DIR / "logs" / "activity.log"
ASSET_TYPES = ["Laptop", "Desktop", "Workstation", "Monitor", "G-Drive", "G-Raid"]
ROLES = ("Super Admin", "Admin", "Asset Manager", "Editor", "Viewer")

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-change-me")
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024

ASSET_COLUMNS = [
    "device_tag", "location", "asset_type", "serial_number", "ip_address", "device_model", "ram_gb",
    "storage_type", "storage_capacity", "action", "status", "username", "system_name", "previous_user_history",
    "purchase_year", "invoice_number", "finance_code", "supplier_name", "asset_value", "warranty_amc",
    "eol", "insurance_status", "insurance_type", "e_code"
]
EMPLOYEE_COLUMNS = ["e_code", "employee_name", "doj", "vertical", "sub_vertical", "content_category", "reporting_manager", "rfds", "employee_status"]
SOFTWARE_COLUMNS = ["software_name", "version", "license_type", "license_key", "number_of_licenses", "sub_vertical", "purchase_date", "registration_email", "status"]
SOFTWARE_ASSIGN_COLUMNS = [
    "desktop_laptop_tag", "e_code", "employee_name", "doj", "vertical", "sub_vertical", "content_category",
    "reporting_manager", "rfds", "location", "software_name", "version", "license_type", "license_key",
    "number_of_licenses", "installation_date", "expiry_date", "vendor_supplier", "purchase_year", "invoice_number",
    "cost", "status", "support_amc", "last_updated_date", "remarks"
]
UNPROCESSED_ASSET_COLUMNS = [
    "device_tag", "e_code", "employee_name", "doj", "vertical", "sub_vertical", "content_category", "reporting_manager",
    "rfds", "location", "asset_type", "action", "status", "username", "system_name", "previous_user_history",
    "serial_number", "ip_address", "device_model", "ram_gb", "storage_type", "storage_capacity", "purchase_year",
    "invoice_number", "finance_code", "supplier_name", "asset_value", "warranty_amc", "eol", "insurance_status", "insurance_type"
]
VENDOR_COLUMNS = ["vendor_name", "contact_person", "email", "phone", "address", "services", "status", "remarks"]
CHECKLIST_COLUMNS = ["checklist_type", "employee_name", "e_code", "vertical", "sub_vertical", "rfds", "payload", "asset_updates", "software_updates", "employee_status", "status"]
NOTIFICATION_TYPES = ["Expiry Alerts", "Assets Approval", "Change", "Asset Removal", "Password Expiry", "Password Changed", "License Expiry"]
DATE_FIELDS = {"doj", "eol", "purchase_date", "installation_date", "expiry_date", "last_updated_date", "email_created_on", "date_of_exit", "form_submission_date", "apply_date", "will_revoke_date", "revoke_date"}
SOPHOS_MACHINE_COLUMNS = ["asset_tag", "hostname", "username", "ticket_id", "group_name", "application_control", "data_loss_prevention", "windows_firewall", "peripheral_control", "threat_protection", "update_management", "web_control", "status", "comment"]
SOPHOS_REQUEST_COLUMNS = ["request_type", "old_hostname", "new_hostname", "hostname", "requester_name", "approver", "engineer", "policy_changed_by", "apply_date", "policy", "ticket_id", "group_name", "duration", "current_status", "will_revoke_date", "day_left_for_revoke", "revoke_date", "access_comment", "revoke_comment", "reason", "payload", "approval_status"]
FEATURES: list[str] = []

SAMPLE_MAP = {
    "sophos-machine-preparation": SOPHOS_REQUEST_COLUMNS,
    "sophos-exit-process": SOPHOS_REQUEST_COLUMNS,
    "sophos-dlp-access": SOPHOS_REQUEST_COLUMNS,
    "sophos-usb-access": SOPHOS_REQUEST_COLUMNS,
    "sophos-remote-access": SOPHOS_REQUEST_COLUMNS,
    "unprocessed-assets": UNPROCESSED_ASSET_COLUMNS,
    "unprocessed-software": SOFTWARE_ASSIGN_COLUMNS,
    "unprocessed-employees": EMPLOYEE_COLUMNS,
    "vendors": VENDOR_COLUMNS,
    "checklist-joining": ["employee_name", "e_code", "vertical", "sub_vertical", "rfds", "machine_asset_tag", "system_serial_number", "model", "ip_address", "operating_system", "office_365_license"],
    "checklist-replacement": ["employee_name", "e_code", "old_assets", "old_softwares", "machine_asset_tag", "system_serial_number", "model", "remarks"],
    "checklist-rebuild": ["employee_name", "e_code", "old_assets", "old_softwares", "machine_asset_tag", "system_serial_number", "model", "remarks"],
    "checklist-exit": ["employee_name", "e_code", "date_of_exit", "form_submission_date", "machine", "data_card", "external_hdd", "host_name", "ad_login_id", "data_handover"],
}

LABELS = {
    "device_tag": "Device Tag", "desktop_laptop_tag": "Desktop/Laptop Tag", "location": "Location", "asset_type": "Asset Type",
    "serial_number": "Serial Number", "ip_address": "IP Address", "device_model": "Device Model", "ram_gb": "RAM GB",
    "storage_type": "Storage Type", "storage_capacity": "Storage Capacity", "action": "Action", "status": "Status",
    "username": "Username", "system_name": "System Name", "previous_user_history": "Previous User History", "purchase_year": "Purchase Year",
    "purchase_date": "Purchase Date",
    "invoice_number": "Invoice Number", "finance_code": "Finance Code", "supplier_name": "Supplier Name", "asset_value": "Asset Value",
    "warranty_amc": "Warranty/AMC", "eol": "End of Life", "insurance_status": "Insurance Status", "insurance_type": "Insurance Type",
    "e_code": "E-Code", "employee_name": "Employee Name", "doj": "Date of Joining", "vertical": "Vertical",
    "sub_vertical": "Sub-Vertical", "content_category": "Content Category", "reporting_manager": "Reporting Manager", "rfds": "RFDs",
    "employee_status": "Employee Status", "software_name": "Software Name", "version": "Version", "license_type": "License Type",
    "license_key": "License Key", "number_of_licenses": "Number of Licenses", "purchase_date": "Purchase Date", "registration_email": "Registration email id",
    "installation_date": "Installation Date", "expiry_date": "Expiry Date", "vendor_supplier": "Vendor/Supplier",
    "cost": "Cost", "support_amc": "Support/AMC", "last_updated_date": "Last Updated Date", "remarks": "Remarks", "vendor_name": "Vendor Name", "contact_person": "Contact Person", "services": "Services", "checklist_type": "Checklist Type", "payload": "Checklist Data", "asset_updates": "Asset Updates", "software_updates": "Software Updates", "asset_tag": "Asset Tag", "hostname": "Hostname", "ticket_id": "Ticket ID", "group_name": "Group", "application_control": "Application Control", "data_loss_prevention": "Data Loss Prevention", "windows_firewall": "Windows Firewall", "peripheral_control": "Peripheral Control", "threat_protection": "Threat Protection", "update_management": "Update Management", "web_control": "Web Control", "request_type": "Request Type", "old_hostname": "Old Hostname", "new_hostname": "New Hostname", "requester_name": "Requester Name", "approver": "Approver", "engineer": "Engineer", "policy_changed_by": "Policy Changed By", "apply_date": "Apply Date", "policy": "Policy", "duration": "Duration", "current_status": "Current Status", "will_revoke_date": "Will Revoke Date", "day_left_for_revoke": "Day Left For Revoke", "revoke_date": "Revoke Date", "access_comment": "Access Comment", "revoke_comment": "Revoke Comment", "reason": "Reason", "approval_status": "Approval Status"
}


def db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def execute(sql: str, params: tuple[Any, ...] = ()) -> sqlite3.Cursor:
    conn = db()
    cur = conn.execute(sql, params)
    conn.commit()
    conn.close()
    return cur


def init_db() -> None:
    schema = f"""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT UNIQUE NOT NULL, password_hash TEXT NOT NULL,
        role TEXT NOT NULL DEFAULT 'Viewer', status TEXT NOT NULL DEFAULT 'Active', disabled INTEGER NOT NULL DEFAULT 0,
        last_login TEXT, last_ip TEXT, full_name TEXT, email TEXT, phone TEXT, department TEXT, theme TEXT DEFAULT 'dark', notification_preferences TEXT DEFAULT '[]', password_changed_at TEXT, features TEXT DEFAULT '[]', created_at TEXT NOT NULL
    );
    CREATE TABLE IF NOT EXISTS employees (id INTEGER PRIMARY KEY AUTOINCREMENT, {cols(EMPLOYEE_COLUMNS)}, created_at TEXT, updated_at TEXT);
    CREATE TABLE IF NOT EXISTS assets (id INTEGER PRIMARY KEY AUTOINCREMENT, {cols(ASSET_COLUMNS)}, approved INTEGER DEFAULT 1, created_at TEXT, updated_at TEXT);
    CREATE TABLE IF NOT EXISTS software (id INTEGER PRIMARY KEY AUTOINCREMENT, {cols(SOFTWARE_COLUMNS)}, used_count INTEGER DEFAULT 0, created_at TEXT, updated_at TEXT);
    CREATE TABLE IF NOT EXISTS software_assignments (id INTEGER PRIMARY KEY AUTOINCREMENT, {cols(SOFTWARE_ASSIGN_COLUMNS)}, created_at TEXT, updated_at TEXT);
    CREATE TABLE IF NOT EXISTS unprocessed_assets (id INTEGER PRIMARY KEY AUTOINCREMENT, {cols(UNPROCESSED_ASSET_COLUMNS)}, maker TEXT, checker TEXT, approval_status TEXT DEFAULT 'Pending', processed INTEGER DEFAULT 0, created_at TEXT, updated_at TEXT);
    CREATE TABLE IF NOT EXISTS unprocessed_software (id INTEGER PRIMARY KEY AUTOINCREMENT, {cols(SOFTWARE_ASSIGN_COLUMNS)}, maker TEXT, checker TEXT, approval_status TEXT DEFAULT 'Pending', processed INTEGER DEFAULT 0, created_at TEXT, updated_at TEXT);
    CREATE TABLE IF NOT EXISTS unprocessed_employees (id INTEGER PRIMARY KEY AUTOINCREMENT, {cols(EMPLOYEE_COLUMNS)}, maker TEXT, checker TEXT, approval_status TEXT DEFAULT 'Pending', processed INTEGER DEFAULT 0, created_at TEXT, updated_at TEXT);
    CREATE TABLE IF NOT EXISTS audit_logs (id INTEGER PRIMARY KEY AUTOINCREMENT, actor TEXT, action TEXT, table_name TEXT, record_id INTEGER, details TEXT, ip TEXT, created_at TEXT NOT NULL);
    CREATE TABLE IF NOT EXISTS login_history (id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT, ip TEXT, status TEXT, created_at TEXT NOT NULL);
    CREATE TABLE IF NOT EXISTS notifications (id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT, type TEXT, message TEXT, is_read INTEGER DEFAULT 0, created_at TEXT NOT NULL);
    CREATE TABLE IF NOT EXISTS email_config (id INTEGER PRIMARY KEY CHECK (id=1), smtp_host TEXT, smtp_port TEXT, smtp_username TEXT, smtp_password TEXT, sender_email TEXT, use_tls INTEGER DEFAULT 1, updated_at TEXT);
    CREATE TABLE IF NOT EXISTS email_logs (id INTEGER PRIMARY KEY AUTOINCREMENT, recipient TEXT, subject TEXT, body TEXT, status TEXT, created_at TEXT NOT NULL);
    CREATE TABLE IF NOT EXISTS vendors (id INTEGER PRIMARY KEY AUTOINCREMENT, {cols(VENDOR_COLUMNS)}, created_at TEXT, updated_at TEXT);
    CREATE TABLE IF NOT EXISTS checklists (id INTEGER PRIMARY KEY AUTOINCREMENT, {cols(CHECKLIST_COLUMNS)}, maker TEXT, checker TEXT, processed INTEGER DEFAULT 0, created_at TEXT, updated_at TEXT);
    CREATE TABLE IF NOT EXISTS checklist_images (id INTEGER PRIMARY KEY AUTOINCREMENT, checklist_id INTEGER, file_path TEXT, original_name TEXT, created_at TEXT NOT NULL);
    CREATE TABLE IF NOT EXISTS sophos_machines (id INTEGER PRIMARY KEY AUTOINCREMENT, {cols(SOPHOS_MACHINE_COLUMNS)}, updated_at TEXT, created_at TEXT);
    CREATE TABLE IF NOT EXISTS sophos_requests (id INTEGER PRIMARY KEY AUTOINCREMENT, {cols(SOPHOS_REQUEST_COLUMNS)}, maker TEXT, checker TEXT, created_at TEXT, updated_at TEXT);
    CREATE TABLE IF NOT EXISTS notification_rules (id INTEGER PRIMARY KEY AUTOINCREMENT, username TEXT, notification_type TEXT, enabled INTEGER DEFAULT 1, email_enabled INTEGER DEFAULT 1, created_at TEXT);
    """
    conn = db()
    conn.executescript(schema)
    conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_assets_device_tag ON assets(device_tag)")
    conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_assets_serial ON assets(serial_number) WHERE serial_number IS NOT NULL AND serial_number != ''")
    conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_employees_ecode ON employees(e_code)")
    conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_software_key ON software(license_key) WHERE license_key IS NOT NULL AND license_key != ''")
    conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_software_name ON software(software_name) WHERE software_name IS NOT NULL AND software_name != ''")
    conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_vendors_name ON vendors(vendor_name) WHERE vendor_name IS NOT NULL AND vendor_name != ''")
    conn.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_assignment_device_license ON software_assignments(desktop_laptop_tag, license_key)")
    ensure_columns(conn)
    valid_roles = tuple(ROLES)
    conn.execute(f"UPDATE users SET role='Editor' WHERE role NOT IN ({','.join('?' for _ in valid_roles)})", valid_roles)
    if not conn.execute("SELECT 1 FROM users WHERE username='admin'").fetchone():
        conn.execute("INSERT INTO users (username,password_hash,role,status,disabled,created_at) VALUES (?,?,?,?,?,?)",
                     ("admin", generate_password_hash("admin123"), "Admin", "Active", 0, now()))
    conn.commit()
    conn.close()


def cols(names: list[str]) -> str:
    return ", ".join(f"{name} TEXT" for name in names)


def ensure_columns(conn: sqlite3.Connection) -> None:
    additions = {
        "users": {"full_name": "TEXT", "email": "TEXT", "phone": "TEXT", "department": "TEXT", "theme": "TEXT DEFAULT 'dark'", "notification_preferences": "TEXT DEFAULT '[]'", "password_changed_at": "TEXT", "features": "TEXT DEFAULT '[]'"},
    }
    for table, columns in additions.items():
        existing = {row[1] for row in conn.execute(f"PRAGMA table_info({table})").fetchall()}
        for column, definition in columns.items():
            if column not in existing:
                conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {definition}")


def now() -> str:
    return datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")


def display_date(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%y"):
        try:
            return datetime.strptime(text[:19] if fmt.endswith("%S") else text[:10], fmt).strftime("%d-%m-%y")
        except ValueError:
            continue
    return text


def current_user() -> sqlite3.Row | None:
    uid = session.get("user_id")
    if not uid:
        return None
    conn = db()
    user = conn.execute("SELECT * FROM users WHERE id=? AND disabled=0", (uid,)).fetchone()
    conn.close()
    return user


def unread_notifications() -> list[sqlite3.Row]:
    username = session.get("username")
    if not username:
        return []
    conn = db()
    rows = conn.execute("SELECT * FROM notifications WHERE username=? AND is_read=0 ORDER BY id DESC LIMIT 10", (username,)).fetchall()
    conn.close()
    return rows


def create_notification(username: str, note_type: str, message: str) -> None:
    execute("INSERT INTO notifications (username,type,message,is_read,created_at) VALUES (?,?,?,?,?)", (username, note_type, message, 0, now()))
    conn = db()
    target = conn.execute("SELECT email, notification_preferences FROM users WHERE username=?", (username,)).fetchone()
    conn.close()
    rule = None
    if target:
        conn = db(); rule = conn.execute("SELECT * FROM notification_rules WHERE username=? AND notification_type=? AND enabled=1", (username, note_type)).fetchone(); conn.close()
    prefs = json.loads(target["notification_preferences"] or "[]") if target else []
    email_allowed = (rule and rule["email_enabled"]) or (not rule and note_type in prefs)
    if target and target["email"] and email_allowed:
        send_email(target["email"], f"EUC Inventory {note_type} Notification", message)


def date_input_value(value: Any) -> str:
    text = str(value or "")[:10]
    try:
        return datetime.strptime(text, "%Y-%m-%d").strftime("%Y-%m-%d")
    except ValueError:
        return ""


def has_feature(user: sqlite3.Row | None, feature: str) -> bool:
    if not user:
        return False
    if user["role"] in ("Admin", "Super Admin"):
        return True
    enabled = json.loads(user["features"] or "[]") if "features" in user.keys() else []
    return feature in enabled


@app.context_processor
def inject_globals() -> dict[str, Any]:
    return {"user": current_user(), "labels": LABELS, "asset_types": ASSET_TYPES, "notification_types": NOTIFICATION_TYPES, "features": FEATURES, "unread_notifications": unread_notifications(), "date_input_value": date_input_value, "display_date": display_date}


def login_required(view):
    @wraps(view)
    def wrapper(*args, **kwargs):
        if not current_user():
            return redirect(url_for("login"))
        return view(*args, **kwargs)
    return wrapper


def roles_required(*roles: str):
    def decorator(view):
        @wraps(view)
        def wrapper(*args, **kwargs):
            user = current_user()
            if not user or (user["role"] != "Super Admin" and user["role"] not in roles):
                flash("You do not have permission for this action.", "danger")
                return redirect(url_for("dashboard"))
            return view(*args, **kwargs)
        return wrapper
    return decorator


def log_action(action: str, table: str = "", record_id: int | None = None, details: Any = None) -> None:
    actor = session.get("username", "system")
    ip = request.headers.get("X-Forwarded-For", request.remote_addr or "") if request else ""
    detail_text = details if isinstance(details, str) else json.dumps(details or {}, default=str)
    LOG_PATH.parent.mkdir(exist_ok=True)
    with LOG_PATH.open("a", encoding="utf-8") as handle:
        handle.write(f"{now()} | {actor} | {ip} | {action} | {table} | {record_id or ''} | {detail_text}\n")
    conn = db()
    conn.execute("INSERT INTO audit_logs (actor,action,table_name,record_id,details,ip,created_at) VALUES (?,?,?,?,?,?,?)",
                 (actor, action, table, record_id, detail_text, ip, now()))
    conn.commit()
    conn.close()


def row_dict(columns: list[str], source: dict[str, Any]) -> dict[str, str]:
    return {col: (source.get(col) or source.get(LABELS.get(col, "")) or "").strip() for col in columns}


def parse_upload(file) -> list[dict[str, str]]:
    filename = (file.filename or "").lower()
    data = file.read()
    rows: list[dict[str, str]] = []
    if filename.endswith(".xlsx"):
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip() for c in next(ws.iter_rows(max_row=1))]
        for record in ws.iter_rows(min_row=2, values_only=True):
            rows.append({headers[i]: str(record[i] or "").strip() for i in range(len(headers))})
    else:
        text = data.decode("utf-8-sig")
        rows = [dict(r) for r in csv.DictReader(io.StringIO(text))]
    return rows



def unique_key_for(table: str) -> str | None:
    return {"assets": "device_tag", "unprocessed_assets": "device_tag", "employees": "e_code", "unprocessed_employees": "e_code", "software": "software_name", "vendors": "vendor_name"}.get(table)


def validate_unique_payload(conn: sqlite3.Connection, table: str, payload: dict[str, str], record_id: int | None = None) -> None:
    key = unique_key_for(table)
    if not key:
        return
    value = str(payload.get(key, "")).strip()
    if not value:
        raise sqlite3.IntegrityError(f"{LABELS.get(key, key)} is mandatory and must be unique")
    params: list[Any] = [value]
    sql = f"SELECT id FROM {table} WHERE {key}=?"
    if table.startswith("unprocessed_"):
        sql += " AND approval_status='Pending'"
    if record_id is not None:
        sql += " AND id!=?"
        params.append(record_id)
    if conn.execute(sql, tuple(params)).fetchone():
        raise sqlite3.IntegrityError(f"Duplicate {LABELS.get(key, key)}: {value}")

def insert_record(table: str, columns: list[str], data: dict[str, str], extra: dict[str, str] | None = None) -> int:
    payload = row_dict(columns, data)
    payload.update(extra or {})
    payload.setdefault("created_at", now())
    payload.setdefault("updated_at", now())
    keys = list(payload.keys())
    placeholders = ",".join("?" for _ in keys)
    conn = db()
    validate_unique_payload(conn, table, payload)
    cur = conn.execute(f"INSERT INTO {table} ({','.join(keys)}) VALUES ({placeholders})", tuple(payload[k] for k in keys))
    conn.commit()
    rid = cur.lastrowid
    conn.close()
    log_action("insert", table, rid, payload)
    return rid


def update_record(table: str, columns: list[str], record_id: int, data: dict[str, str]) -> None:
    payload = row_dict(columns, data)
    payload["updated_at"] = now()
    assignments = ",".join(f"{k}=?" for k in payload)
    conn = db()
    validate_unique_payload(conn, table, payload, record_id)
    conn.execute(f"UPDATE {table} SET {assignments} WHERE id=?", tuple(payload.values()) + (record_id,))
    conn.commit(); conn.close()
    log_action("update", table, record_id, payload)


def approve_asset(record_id: int) -> None:
    conn = db()
    row = conn.execute("SELECT * FROM unprocessed_assets WHERE id=?", (record_id,)).fetchone()
    if not row:
        raise ValueError("Asset request not found")
    data = dict(row)
    employee = row_dict(EMPLOYEE_COLUMNS, data)
    if employee["e_code"]:
        upsert(conn, "employees", EMPLOYEE_COLUMNS, employee, "e_code")
    asset = row_dict(ASSET_COLUMNS, data)
    upsert(conn, "assets", ASSET_COLUMNS, asset, "device_tag")
    conn.execute("UPDATE unprocessed_assets SET approval_status='Approved', processed=1, checker=?, updated_at=? WHERE id=?", (session.get("username"), now(), record_id))
    conn.commit(); conn.close()
    log_action("approve", "unprocessed_assets", record_id, {"device_tag": data.get("device_tag")})


def approve_software(record_id: int) -> None:
    conn = db()
    row = conn.execute("SELECT * FROM unprocessed_software WHERE id=?", (record_id,)).fetchone()
    if not row:
        raise ValueError("Software request not found")
    data = dict(row)
    software = row_dict(SOFTWARE_COLUMNS, data)
    software["purchase_date"] = data.get("installation_date", "")
    upsert(conn, "software", SOFTWARE_COLUMNS, software, "software_name")
    assignment = row_dict(SOFTWARE_ASSIGN_COLUMNS, data)
    upsert(conn, "software_assignments", SOFTWARE_ASSIGN_COLUMNS, assignment, "desktop_laptop_tag", extra_unique="license_key")
    conn.execute("UPDATE software SET used_count=(SELECT COUNT(*) FROM software_assignments WHERE software_assignments.license_key=software.license_key)")
    conn.execute("UPDATE unprocessed_software SET approval_status='Approved', processed=1, checker=?, updated_at=? WHERE id=?", (session.get("username"), now(), record_id))
    conn.commit(); conn.close()
    log_action("approve", "unprocessed_software", record_id, {"license_key": data.get("license_key")})


def approve_employee(record_id: int) -> None:
    conn = db()
    row = conn.execute("SELECT * FROM unprocessed_employees WHERE id=?", (record_id,)).fetchone()
    if not row:
        raise ValueError("Employee request not found")
    upsert(conn, "employees", EMPLOYEE_COLUMNS, row_dict(EMPLOYEE_COLUMNS, dict(row)), "e_code")
    conn.execute("UPDATE unprocessed_employees SET approval_status='Approved', processed=1, checker=?, updated_at=? WHERE id=?", (session.get("username"), now(), record_id))
    conn.commit(); conn.close()
    log_action("approve", "unprocessed_employees", record_id, {"e_code": row["e_code"]})


def upsert(conn: sqlite3.Connection, table: str, columns: list[str], payload: dict[str, str], unique: str, extra_unique: str | None = None) -> None:
    payload["updated_at"] = now()
    where = f"{unique}=?"
    params = [payload.get(unique, "")]
    if extra_unique:
        where += f" AND {extra_unique}=?"
        params.append(payload.get(extra_unique, ""))
    existing = conn.execute(f"SELECT id FROM {table} WHERE {where}", params).fetchone()
    if existing:
        assignments = ",".join(f"{k}=?" for k in payload)
        conn.execute(f"UPDATE {table} SET {assignments} WHERE id=?", tuple(payload.values()) + (existing["id"],))
    else:
        payload["created_at"] = now()
        keys = list(payload.keys())
        conn.execute(f"INSERT INTO {table} ({','.join(keys)}) VALUES ({','.join('?' for _ in keys)})", tuple(payload.values()))



def reference_context() -> dict[str, Any]:
    conn = db()
    assets = [dict(r) for r in conn.execute("SELECT * FROM assets ORDER BY device_tag LIMIT 1000").fetchall()]
    employees = [dict(r) for r in conn.execute("SELECT * FROM employees ORDER BY employee_name LIMIT 1000").fetchall()]
    software = [dict(r) for r in conn.execute("SELECT * FROM software ORDER BY software_name LIMIT 1000").fetchall()]
    assignments = [dict(r) for r in conn.execute("SELECT * FROM software_assignments ORDER BY updated_at DESC LIMIT 1000").fetchall()]
    vendors = [dict(r) for r in conn.execute("SELECT * FROM vendors ORDER BY vendor_name LIMIT 1000").fetchall()]
    asset_requests = [dict(r) for r in conn.execute("SELECT * FROM unprocessed_assets ORDER BY updated_at DESC LIMIT 1000").fetchall()]
    conn.close()
    return {
        "ref_assets": assets,
        "ref_employees": employees,
        "ref_software": software,
        "ref_assignments": assignments,
        "ref_vendors": vendors,
        "asset_history_rows": asset_requests,
    }


def compact_form_data(form) -> dict[str, str]:
    return {k: v for k, v in form.to_dict().items() if not k.startswith("enable_") and k != "form_type" and str(v).strip()}

def rows_for(table: str, order: str = "id DESC", limit: int | None = None, where: str = "1=1", params: tuple[Any, ...] = ()) -> list[sqlite3.Row]:
    conn = db()
    sql = f"SELECT * FROM {table} WHERE {where} ORDER BY {order}"
    if limit:
        sql += f" LIMIT {limit}"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return rows


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"]
        conn = db()
        user = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()
        ok = user and not user["disabled"] and user["status"] == "Active" and check_password_hash(user["password_hash"], password)
        ip = request.remote_addr or ""
        conn.execute("INSERT INTO login_history (username,ip,status,created_at) VALUES (?,?,?,?)", (username, ip, "Success" if ok else "Failed", now()))
        if ok:
            conn.execute("UPDATE users SET last_login=?, last_ip=? WHERE id=?", (now(), ip, user["id"]))
            session.update(user_id=user["id"], username=user["username"], role=user["role"])
            conn.commit(); conn.close(); log_action("login", "users", user["id"])
            return redirect(url_for("dashboard"))
        conn.commit(); conn.close(); flash("Invalid credentials or disabled user.", "danger")
    return render_template("login.html")


@app.route("/forgot-password", methods=["GET", "POST"])
def forgot_password():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        conn = db(); user = conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone(); conn.close()
        if user and user["email"]:
            temp = "Temp@" + datetime.utcnow().strftime("%H%M%S")
            execute("UPDATE users SET password_hash=?, password_changed_at=? WHERE id=?", (generate_password_hash(temp), now(), user["id"]))
            send_email(user["email"], "EUC Inventory password reset", f"Temporary password: {temp}\nPlease login and change it immediately.")
        flash("If the account has an email configured, a temporary password has been sent.", "success")
        return redirect(url_for("login"))
    return render_template("forgot_password.html")


@app.route("/logout")
def logout():
    log_action("logout")
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
@login_required
def dashboard():
    conn = db()
    asset_total = conn.execute("SELECT COUNT(*) c FROM assets").fetchone()["c"]
    employee_total = conn.execute("SELECT COUNT(*) c FROM employees").fetchone()["c"]
    software_total = conn.execute("SELECT COUNT(*) c FROM software").fetchone()["c"]
    active_users = conn.execute("SELECT COUNT(*) c FROM users WHERE disabled=0 AND status='Active'").fetchone()["c"]
    total_users = conn.execute("SELECT COUNT(*) c FROM users").fetchone()["c"]
    counts = {t: conn.execute("SELECT COUNT(*) c FROM assets WHERE asset_type=?", (t,)).fetchone()["c"] for t in ASSET_TYPES}
    approved = conn.execute("SELECT COUNT(*) c FROM assets WHERE approved=1").fetchone()["c"]
    pending = conn.execute("SELECT COUNT(*) c FROM unprocessed_assets WHERE approval_status='Pending'").fetchone()["c"]
    software = conn.execute("SELECT *, MAX(CAST(number_of_licenses AS INTEGER)-used_count,0) free_count FROM software ORDER BY software_name").fetchall()
    unapproved = conn.execute("SELECT asset_type, COUNT(*) c FROM unprocessed_assets WHERE approval_status='Pending' GROUP BY asset_type").fetchall()
    sophos_counts = {row["request_type"] or "Unknown": row["c"] for row in conn.execute("SELECT request_type, COUNT(*) c FROM sophos_requests GROUP BY request_type").fetchall()}
    sophos_pending = conn.execute("SELECT COUNT(*) c FROM sophos_requests WHERE approval_status='Pending'").fetchone()["c"]
    sophos_approved = conn.execute("SELECT COUNT(*) c FROM sophos_requests WHERE approval_status='Approved'").fetchone()["c"]
    task_counts = {"Asset Approval": pending, "Checklist Approval": conn.execute("SELECT COUNT(*) c FROM checklists WHERE status='Pending'").fetchone()["c"], "Sophos Approval": sophos_pending}
    software_pie = {row["status"] or "Unknown": row["c"] for row in conn.execute("SELECT status, COUNT(*) c FROM software GROUP BY status").fetchall()}
    logs = conn.execute("SELECT * FROM audit_logs ORDER BY id DESC LIMIT 200").fetchall()
    logins = conn.execute("SELECT * FROM login_history ORDER BY id DESC LIMIT 200").fetchall()
    conn.close()
    return render_template("dashboard.html", asset_total=asset_total, employee_total=employee_total, software_total=software_total,
                           active_users=active_users, total_users=total_users, counts=counts, approved=approved, pending=pending,
                           software=software, unapproved=unapproved, logs=logs, logins=logins, sophos_counts=sophos_counts,
                           sophos_pending=sophos_pending, sophos_approved=sophos_approved, task_counts=task_counts, software_pie=software_pie)


@app.route("/inventory/<kind>", methods=["GET", "POST"])
@login_required
def inventory(kind: str):
    config = {
        "assets": ("Assets Inventory", "assets", ASSET_COLUMNS, ["Super Admin", "Admin", "Asset Manager"]),
        "software": ("Licensed Software", "software", SOFTWARE_COLUMNS, ["Super Admin", "Admin", "Asset Manager"]),
        "employees": ("Employee List", "employees", EMPLOYEE_COLUMNS, ["Super Admin", "Admin", "Asset Manager"]),
        "unprocessed-assets": ("Unprocessed Assets", "unprocessed_assets", UNPROCESSED_ASSET_COLUMNS, ["Super Admin", "Admin", "Asset Manager", "Editor"]),
        "unprocessed-software": ("Unprocessed Software", "unprocessed_software", SOFTWARE_ASSIGN_COLUMNS, ["Super Admin", "Admin", "Asset Manager", "Editor"]),
        "unprocessed-employees": ("Unprocessed Employees", "unprocessed_employees", EMPLOYEE_COLUMNS, ["Super Admin", "Admin", "Asset Manager", "Editor"]),
    }.get(kind)
    if not config:
        return redirect(url_for("dashboard"))
    title, table, columns, writers = config
    user = current_user()
    if request.method == "POST" and user["role"] in writers:
        extra = {}
        if table.startswith("unprocessed"):
            extra = {"maker": user["username"], "approval_status": "Pending", "processed": "0"}
        try:
            if kind == "software" and request.form.get("form_type") == "software_allocation":
                selected_keys = request.form.getlist("license_keys") or [request.form.get("license_key", "")]
                saved = 0
                base = compact_form_data(request.form)
                for key in selected_keys:
                    if not key.strip():
                        continue
                    payload = dict(base)
                    payload["license_key"] = key.strip()
                    match = next((s for s in reference_context()["ref_software"] if s.get("license_key") == key.strip()), None)
                    if match:
                        total = int(match.get("number_of_licenses") or 0)
                        used = int(match.get("used_count") or 0)
                        if total and used >= total:
                            flash(f"No licenses left for {match.get('software_name')}", "warning")
                            continue
                        payload.update({k: payload.get(k) or match.get(k, "") for k in ("software_name", "version", "license_type", "number_of_licenses", "status")})
                    insert_record("unprocessed_software", SOFTWARE_ASSIGN_COLUMNS, payload, {"maker": user["username"], "approval_status": "Pending", "processed": "0"})
                    saved += 1
                flash(f"Submitted {saved} software allocation request(s) for approval.", "success")
            else:
                rid = insert_record(table, columns, compact_form_data(request.form), extra)
                flash(f"Saved record #{rid}.", "success")
        except sqlite3.IntegrityError as exc:
            flash(f"Duplicate entry blocked: {exc}", "danger")
        return redirect(url_for("inventory", kind=kind))
    search = request.args.get("q", "").strip()
    where, params = ("approval_status='Pending'", ()) if table.startswith("unprocessed") else ("1=1", ())
    if search:
        like = f"%{search}%"
        search_where = " OR ".join(f"{c} LIKE ?" for c in columns)
        where = f"({where}) AND ({search_where})"
        params = params + tuple([like] * len(columns))
    rows = rows_for(table, where=where, params=params)
    return render_template("inventory.html", kind=kind, title=title, table=table, columns=columns, rows=rows, can_write=user["role"] in writers, **reference_context())


def validate_upload(table: str, columns: list[str], records: list[dict[str, str]]) -> list[str]:
    errors: list[str] = []
    headers = set(records[0].keys()) if records else set()
    missing = [LABELS.get(c, c) for c in columns if c not in headers and LABELS.get(c, "") not in headers]
    if missing:
        errors.append("Missing columns: " + ", ".join(missing[:12]))
    conn = db()
    for idx, record in enumerate(records, start=2):
        normalized = row_dict(columns, record)
        if table == "unprocessed_assets":
            tag, serial = normalized.get("device_tag", ""), normalized.get("serial_number", "")
            if tag and conn.execute("SELECT 1 FROM assets WHERE device_tag=? UNION SELECT 1 FROM unprocessed_assets WHERE device_tag=? AND approval_status='Pending'", (tag, tag)).fetchone():
                errors.append(f"Row {idx}: duplicate pending/approved device tag {tag}")
            if serial and conn.execute("SELECT 1 FROM assets WHERE serial_number=? UNION SELECT 1 FROM unprocessed_assets WHERE serial_number=? AND approval_status='Pending'", (serial, serial)).fetchone():
                errors.append(f"Row {idx}: duplicate pending/approved serial number {serial}")
        elif table == "unprocessed_software":
            key, tag = normalized.get("license_key", ""), normalized.get("desktop_laptop_tag", "")
            if key and tag and conn.execute("SELECT 1 FROM software_assignments WHERE desktop_laptop_tag=? AND license_key=? UNION SELECT 1 FROM unprocessed_software WHERE desktop_laptop_tag=? AND license_key=? AND approval_status='Pending'", (tag, key, tag, key)).fetchone():
                errors.append(f"Row {idx}: duplicate software allocation for {tag}/{key}")
        elif table == "unprocessed_employees":
            ecode = normalized.get("e_code", "")
            if ecode and conn.execute("SELECT 1 FROM employees WHERE e_code=? UNION SELECT 1 FROM unprocessed_employees WHERE e_code=? AND approval_status='Pending'", (ecode, ecode)).fetchone():
                errors.append(f"Row {idx}: duplicate pending/approved employee {ecode}")
    conn.close()
    return errors


@app.route("/inventory/<kind>/upload", methods=["POST"])
@login_required
@roles_required("Admin", "Asset Manager", "Editor")
def upload(kind: str):
    maps = {"unprocessed-assets": ("unprocessed_assets", UNPROCESSED_ASSET_COLUMNS), "unprocessed-software": ("unprocessed_software", SOFTWARE_ASSIGN_COLUMNS), "unprocessed-employees": ("unprocessed_employees", EMPLOYEE_COLUMNS)}
    if kind not in maps:
        return redirect(url_for("inventory", kind=kind))
    table, columns = maps[kind]
    try:
        records = parse_upload(request.files["file"])
        errors = validate_upload(table, columns, records)
        if errors:
            for error in errors[:10]:
                flash(error, "danger")
            if len(errors) > 10:
                flash(f"{len(errors) - 10} more validation errors hidden. Correct the file and upload again.", "danger")
            log_action("upload_validation_failed", table, details={"errors": errors[:25]})
            return redirect(url_for("inventory", kind=kind))
        for record in records:
            insert_record(table, columns, record, {"maker": session["username"], "approval_status": "Pending", "processed": "0"})
        create_notification(session["username"], "Assets Approval", f"{len(records)} {kind} records submitted for approval")
        flash(f"Validated and uploaded {len(records)} pending records.", "success")
    except Exception as exc:
        flash(f"Upload failed: {exc}", "danger")
    return redirect(url_for("inventory", kind=kind))


@app.route("/record/<table>/<int:record_id>/edit", methods=["GET", "POST"])
@login_required
def edit_record(table: str, record_id: int):
    allowed = {"assets": ASSET_COLUMNS, "software": SOFTWARE_COLUMNS, "employees": EMPLOYEE_COLUMNS, "unprocessed_assets": UNPROCESSED_ASSET_COLUMNS, "unprocessed_software": SOFTWARE_ASSIGN_COLUMNS, "unprocessed_employees": EMPLOYEE_COLUMNS}
    if table not in allowed:
        return redirect(url_for("dashboard"))
    if current_user()["role"] == "Viewer":
        flash("Viewers cannot edit records.", "danger"); return redirect(url_for("dashboard"))
    columns = allowed[table]
    conn = db(); row = conn.execute(f"SELECT * FROM {table} WHERE id=?", (record_id,)).fetchone(); conn.close()
    if request.method == "POST":
        try:
            if table in ("assets", "software", "employees"):
                staging = {"assets": ("unprocessed_assets", UNPROCESSED_ASSET_COLUMNS), "software": ("unprocessed_software", SOFTWARE_ASSIGN_COLUMNS), "employees": ("unprocessed_employees", EMPLOYEE_COLUMNS)}[table]
                source = dict(row or {})
                source.update(request.form.to_dict())
                source["previous_user_history"] = (source.get("previous_user_history", "") + f"\nChange requested by {session.get('username')} on {now()} from live {table} #{record_id}").strip()
                insert_record(staging[0], staging[1], source, {"maker": session["username"], "approval_status": "Pending", "processed": "0"})
                create_notification(session["username"], "Change", f"Live {table} edit submitted for approval")
                flash("Live table changes were submitted to unprocessed approvals. Final data will update after approval.", "success")
                return redirect(url_for("inventory", kind={"assets":"unprocessed-assets","software":"unprocessed-software","employees":"unprocessed-employees"}[table]))
            update_record(table, columns, record_id, request.form)
            flash("Record updated.", "success")
        except sqlite3.IntegrityError as exc:
            flash(f"Duplicate entry blocked: {exc}", "danger")
    return render_template("edit.html", table=table, row=row, columns=columns, back_url=request.referrer or url_for("dashboard"), **reference_context())


@app.route("/approve/<kind>/<int:record_id>", methods=["POST"])
@login_required
@roles_required("Admin", "Asset Manager")
def approve(kind: str, record_id: int):
    try:
        {"asset": approve_asset, "software": approve_software, "employee": approve_employee}[kind](record_id)
        create_notification(session["username"], "Change", "Approval completed and final inventory updated")
        flash("Request approved and posted to final inventory.", "success")
    except (KeyError, sqlite3.IntegrityError, ValueError) as exc:
        flash(f"Approval failed: {exc}", "danger")
    target = {"asset": "unprocessed-assets", "software": "unprocessed-software", "employee": "unprocessed-employees"}.get(kind, "assets")
    return redirect(url_for("inventory", kind=target))


@app.route("/reject/<table>/<int:record_id>", methods=["POST"])
@login_required
@roles_required("Admin", "Asset Manager")
def reject(table: str, record_id: int):
    if table not in ("unprocessed_assets", "unprocessed_software", "unprocessed_employees"):
        return redirect(url_for("dashboard"))
    execute(f"UPDATE {table} SET approval_status='Rejected', checker=?, updated_at=? WHERE id=?", (session["username"], now(), record_id))
    log_action("reject", table, record_id)
    create_notification(session["username"], "Assets Approval", f"{table} request #{record_id} rejected")
    flash("Request rejected.", "warning")
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/reports/<name>")
@login_required
def report(name: str):
    start, end, asset_filter, software_filter = request.args.get("start", ""), request.args.get("end", ""), request.args.get("asset", ""), request.args.get("software", "")
    conn = db()
    if name == "inventory":
        columns = ["device_tag"] + EMPLOYEE_COLUMNS + [c for c in ASSET_COLUMNS if c != "e_code"]
        sql = "SELECT a.device_tag,e.*,a.location,a.asset_type,a.action,a.status,a.username,a.system_name,a.previous_user_history,a.serial_number,a.ip_address,a.device_model,a.ram_gb,a.storage_type,a.storage_capacity,a.purchase_year,a.invoice_number,a.finance_code,a.supplier_name,a.asset_value,a.warranty_amc,a.eol,a.insurance_status,a.insurance_type FROM assets a LEFT JOIN employees e ON a.e_code=e.e_code WHERE 1=1"
        params=[]
        if asset_filter: sql += " AND a.asset_type=?"; params.append(asset_filter)
        if start: sql += " AND date(a.created_at)>=date(?)"; params.append(start)
        if end: sql += " AND date(a.created_at)<=date(?)"; params.append(end)
        rows = conn.execute(sql, params).fetchall(); title = "Inventory Report"
    elif name == "software":
        columns = SOFTWARE_ASSIGN_COLUMNS
        sql = "SELECT * FROM software_assignments WHERE 1=1"; params=[]
        if software_filter: sql += " AND software_name LIKE ?"; params.append(f"%{software_filter}%")
        if start: sql += " AND date(created_at)>=date(?)"; params.append(start)
        if end: sql += " AND date(created_at)<=date(?)"; params.append(end)
        rows = conn.execute(sql, params).fetchall(); title = "Software Report"
    elif name == "allotted":
        columns = ["employee_name", "e_code", "device_tag", "asset_type", "created_at", "previous_user_history"]
        sql = "SELECT e.employee_name,a.e_code,a.device_tag,a.asset_type,a.created_at,a.previous_user_history FROM assets a LEFT JOIN employees e ON a.e_code=e.e_code WHERE a.e_code IS NOT NULL AND a.e_code!=''"
        params=[]
        if asset_filter: sql += " AND a.asset_type=?"; params.append(asset_filter)
        if start: sql += " AND date(a.created_at)>=date(?)"; params.append(start)
        if end: sql += " AND date(a.created_at)<=date(?)"; params.append(end)
        rows = conn.execute(sql, params).fetchall(); title = "Employee Allotted Assets Report"
    else:
        columns = ASSET_COLUMNS
        sql = "SELECT * FROM assets WHERE (e_code IS NULL OR e_code='' OR username IS NULL OR username='')"
        params=[]
        if asset_filter: sql += " AND asset_type=?"; params.append(asset_filter)
        if start: sql += " AND date(created_at)>=date(?)"; params.append(start)
        if end: sql += " AND date(created_at)<=date(?)"; params.append(end)
        rows = conn.execute(sql, params).fetchall(); title = "Unallocated Assets"
    conn.close()
    if request.args.get("download"):
        return csv_response(title, columns, rows)
    return render_template("report.html", name=name, title=title, columns=columns, rows=rows, start=start, end=end, asset_filter=asset_filter, software_filter=software_filter)


def csv_response(title: str, columns: list[str], rows: list[sqlite3.Row]) -> Response:
    output = io.StringIO(); writer = csv.writer(output)
    writer.writerow([LABELS.get(c, c.replace("_", " ").title()) for c in columns])
    for row in rows:
        writer.writerow([row[c] if c in row.keys() else "" for c in columns])
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": f"attachment; filename={title.lower().replace(' ', '_')}.csv"})


@app.route("/admin", methods=["GET", "POST"])
@login_required
@roles_required("Admin")
def admin():
    if request.method == "POST":
        action = request.form.get("action")
        try:
            if action == "create_user":
                execute("INSERT INTO users (username,password_hash,role,status,disabled,created_at) VALUES (?,?,?,?,?,?)",
                        (request.form["username"].strip(), generate_password_hash(request.form["password"]), request.form["role"], "Active", 0, now()))
                log_action("create_user", "users", details={"username": request.form["username"], "role": request.form["role"]})
            elif action == "update_user":
                disabled = 1 if request.form.get("disabled") == "yes" else 0
                params: list[Any] = [request.form["role"], disabled, "Disabled" if disabled else "Active", "[]"]
                sql = "UPDATE users SET role=?, disabled=?, status=?, features=?"
                if request.form.get("new_password"):
                    sql += ", password_hash=?"; params.append(generate_password_hash(request.form["new_password"]))
                sql += " WHERE id=?"; params.append(request.form["user_id"])
                execute(sql, tuple(params)); log_action("update_user", "users", int(request.form["user_id"]))
            flash("User control updated.", "success")
        except sqlite3.IntegrityError as exc:
            flash(f"User update failed: {exc}", "danger")
        return redirect(url_for("admin"))
    conn = db()
    users = conn.execute("SELECT id,username,role,status,disabled,last_login,last_ip,features,created_at FROM users ORDER BY username").fetchall()
    logs = conn.execute("SELECT * FROM audit_logs ORDER BY id DESC LIMIT 500").fetchall()
    assets = conn.execute("SELECT * FROM unprocessed_assets ORDER BY id DESC").fetchall()
    conn.close()
    return render_template("admin.html", users=users, roles=ROLES, logs=logs, assets=assets)


@app.route("/sample/<kind>")
@login_required
def sample_csv(kind: str):
    columns = SAMPLE_MAP.get(kind)
    if not columns:
        return redirect(url_for("dashboard"))
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([LABELS.get(c, c.replace("_", " ").title()) for c in columns])
    writer.writerow(["sample" if c not in DATE_FIELDS else date.today().isoformat() for c in columns])
    return Response(output.getvalue(), mimetype="text/csv", headers={"Content-Disposition": f"attachment; filename={kind}_sample.csv"})


@app.route("/delete/<table>/<int:record_id>", methods=["POST"])
@login_required
@roles_required("Admin", "Asset Manager", "Editor")
def delete_record(table: str, record_id: int):
    allowed = {"unprocessed_assets", "unprocessed_software", "unprocessed_employees", "vendors"}
    if table not in allowed:
        return redirect(url_for("dashboard"))
    if table.startswith("unprocessed"):
        execute(f"DELETE FROM {table} WHERE id=? AND approval_status!='Approved'", (record_id,))
    else:
        execute(f"DELETE FROM {table} WHERE id=?", (record_id,))
    log_action("delete", table, record_id)
    flash("Row deleted.", "warning")
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/profile", methods=["GET", "POST"])
@login_required
def profile():
    user = current_user()
    if request.method == "POST":
        theme = request.form.get("theme", "dark")
        params: list[Any] = [request.form.get("full_name", ""), request.form.get("email", ""), request.form.get("phone", ""), request.form.get("department", ""), theme]
        sql = "UPDATE users SET full_name=?, email=?, phone=?, department=?, theme=?"
        if request.form.get("new_password"):
            if not check_password_hash(user["password_hash"], request.form.get("current_password", "")):
                flash("Current password is incorrect.", "danger")
                return redirect(url_for("profile"))
            sql += ", password_hash=?, password_changed_at=?"
            params.extend([generate_password_hash(request.form["new_password"]), now()])
        sql += " WHERE id=?"
        params.append(user["id"])
        execute(sql, tuple(params))
        session["theme"] = theme
        log_action("profile_update", "users", user["id"], {"theme": theme})
        flash("Profile updated.", "success")
        return redirect(url_for("profile"))
    prefs = json.loads(user["notification_preferences"] or "[]") if user else []
    conn = db()
    notes = conn.execute("SELECT * FROM notifications WHERE username=? ORDER BY id DESC LIMIT 100", (user["username"],)).fetchall()
    conn.close()
    return render_template("profile.html", prefs=prefs, notes=notes)


@app.route("/notifications/read", methods=["POST"])
@login_required
def read_notifications():
    execute("UPDATE notifications SET is_read=1 WHERE username=?", (session["username"],))
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/admin/logs")
@login_required
@roles_required("Admin")
def admin_logs():
    q = request.args.get("q", "")
    conn = db()
    like = f"%{q}%"
    logs = conn.execute("SELECT * FROM audit_logs WHERE actor LIKE ? OR action LIKE ? OR table_name LIKE ? OR details LIKE ? ORDER BY id DESC LIMIT 500", (like, like, like, like)).fetchall()
    assets = conn.execute("SELECT id,device_tag,asset_type,approval_status,maker,checker,updated_at FROM unprocessed_assets ORDER BY id DESC LIMIT 300").fetchall()
    emails = conn.execute("SELECT * FROM email_logs ORDER BY id DESC LIMIT 300").fetchall()
    conn.close()
    return render_template("admin_logs.html", logs=logs, assets=assets, emails=emails, q=q)


@app.route("/admin/db", methods=["GET", "POST"])
@login_required
@roles_required("Admin")
def admin_db():
    result, columns = [], []
    predefined = {
        "pending_assets": "SELECT id, device_tag, asset_type, maker, created_at FROM unprocessed_assets WHERE approval_status='Pending' ORDER BY id DESC LIMIT 100",
        "expiring_licenses": "SELECT software_name, license_key, expiry_date, status FROM software_assignments WHERE expiry_date!='' ORDER BY expiry_date LIMIT 100",
        "inactive_employees": "SELECT e_code, employee_name, employee_status FROM employees WHERE employee_status!='Active' LIMIT 100",
        "unallocated_assets": "SELECT device_tag, asset_type, status, location FROM assets WHERE e_code IS NULL OR e_code='' LIMIT 100",
        "pending_sophos": "SELECT id, request_type, hostname, ticket_id, maker FROM sophos_requests WHERE approval_status='Pending' ORDER BY id DESC LIMIT 100",
        "sophos_expiring_30_days": "SELECT hostname, request_type, policy, will_revoke_date, day_left_for_revoke FROM sophos_requests WHERE CAST(day_left_for_revoke AS INTEGER) BETWEEN 0 AND 30 ORDER BY CAST(day_left_for_revoke AS INTEGER) LIMIT 100",
    }
    if request.method == "POST":
        action = request.form.get("action")
        if action == "backup":
            backup_dir = BASE_DIR / "backups"; backup_dir.mkdir(exist_ok=True)
            target = backup_dir / f"inventory_backup_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.db"
            shutil.copy2(DB_PATH, target)
            log_action("db_backup", "database", details=str(target))
            flash(f"Backup created: {target.name}", "success")
        elif action == "restore" and request.files.get("db_file"):
            request.files["db_file"].save(DB_PATH)
            log_action("db_restore", "database")
            flash("Database restored. Please restart the app if needed.", "warning")
        else:
            query = predefined.get(request.form.get("predefined"), request.form.get("query", "").strip())
            if not query.lower().startswith("select"):
                flash("Only SELECT queries are allowed from the web console.", "danger")
            else:
                conn = db(); cur = conn.execute(query); rows = cur.fetchall(); conn.close()
                columns = [d[0] for d in cur.description or []]
                result = rows
                log_action("db_query", "database", details=query)
    conn = db()
    table_counts = []
    for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name").fetchall():
        table_counts.append({"name": row["name"], "count": conn.execute(f"SELECT COUNT(*) c FROM {row['name']}").fetchone()["c"]})
    conn.close()
    return render_template("admin_db.html", predefined=predefined, result=result, columns=columns, table_counts=table_counts)


@app.route("/admin/email", methods=["GET", "POST"])
@login_required
@roles_required("Admin")
def admin_email():
    if request.method == "POST":
        if request.form.get("action") == "save":
            execute("INSERT OR REPLACE INTO email_config (id,smtp_host,smtp_port,smtp_username,smtp_password,sender_email,use_tls,updated_at) VALUES (1?,?,?,?,?,?,?)",
                    (request.form.get("smtp_host", ""), request.form.get("smtp_port", "587"), request.form.get("smtp_username", ""), request.form.get("smtp_password", ""), request.form.get("sender_email", ""), 1 if request.form.get("use_tls") else 0, now()))
            flash("Email server configuration saved.", "success")
        elif request.form.get("action") == "test":
            status = send_email(request.form.get("test_recipient", ""), "EUC Inventory test email", "This is a test notification from EUC Inventory.")
            flash(status, "success" if status.startswith("Sent") else "danger")
        return redirect(url_for("admin_email"))
    conn = db()
    config = conn.execute("SELECT * FROM email_config WHERE id=1").fetchone()
    logs = conn.execute("SELECT * FROM email_logs ORDER BY id DESC LIMIT 200").fetchall()
    conn.close()
    return render_template("admin_email.html", config=config, logs=logs)


def send_email(recipient: str, subject: str, body: str) -> str:
    conn = db(); config = conn.execute("SELECT * FROM email_config WHERE id=1").fetchone(); conn.close()
    if not config or not recipient:
        status = "Email configuration or recipient missing"
    else:
        message = f"From: {config['sender_email']}\r\nTo: {recipient}\r\nSubject: {subject}\r\n\r\n{body}"
        try:
            server = smtplib.SMTP(config["smtp_host"], int(config["smtp_port"] or 587), timeout=10)
            if config["use_tls"]:
                server.starttls()
            if config["smtp_username"]:
                server.login(config["smtp_username"], config["smtp_password"])
            server.sendmail(config["sender_email"], [recipient], message)
            server.quit()
            status = "Sent test email"
        except Exception as exc:
            status = f"Email failed: {exc}"
    execute("INSERT INTO email_logs (recipient,subject,body,status,created_at) VALUES (?,?,?,?,?)", (recipient, subject, body, status, now()))
    return status


@app.route("/inventory/vendors", methods=["GET", "POST"])
@login_required
def vendors():
    if request.method == "POST" and current_user()["role"] in ("Admin", "Editor"):
        if request.files.get("file") and request.files["file"].filename:
            records = parse_upload(request.files["file"])
            for record in records:
                insert_record("vendors", VENDOR_COLUMNS, record)
            flash(f"Uploaded {len(records)} vendors.", "success")
        else:
            insert_record("vendors", VENDOR_COLUMNS, request.form)
            flash("Vendor saved.", "success")
        return redirect(url_for("vendors"))
    search = request.args.get("q", "")
    where, params = "1=1", ()
    if search:
        like = f"%{search}%"; where = " OR ".join(f"{c} LIKE ?" for c in VENDOR_COLUMNS); params = tuple([like] * len(VENDOR_COLUMNS))
    conn = db()
    rows = conn.execute(f"SELECT v.*, (SELECT COUNT(*) FROM assets a WHERE a.supplier_name=v.vendor_name) asset_count, (SELECT COUNT(*) FROM software_assignments s WHERE s.vendor_supplier=v.vendor_name) software_count FROM vendors v WHERE {where} ORDER BY id DESC", params).fetchall()
    assets = conn.execute("SELECT supplier_name, COUNT(*) c FROM assets WHERE supplier_name!='' GROUP BY supplier_name").fetchall()
    software = conn.execute("SELECT vendor_supplier, COUNT(*) c FROM software_assignments WHERE vendor_supplier!='' GROUP BY vendor_supplier").fetchall()
    vendor_assets = [dict(r) for r in conn.execute("SELECT device_tag, system_name, supplier_name, invoice_number, purchase_year, asset_value FROM assets WHERE supplier_name!='' ORDER BY supplier_name, device_tag").fetchall()]
    vendor_software = [dict(r) for r in conn.execute("SELECT software_name, license_key, desktop_laptop_tag, employee_name, vendor_supplier, invoice_number, purchase_year, cost FROM software_assignments WHERE vendor_supplier!='' ORDER BY vendor_supplier, software_name").fetchall()]
    conn.close()
    return render_template("vendors.html", rows=rows, columns=VENDOR_COLUMNS, assets=assets, software=software, vendor_assets=vendor_assets, vendor_software=vendor_software)


def asset_context(ctype: str) -> dict[str, Any]:
    conn = db()
    pending = conn.execute("SELECT * FROM checklists WHERE status='Pending' AND checklist_type=? ORDER BY id DESC LIMIT 200", (ctype,)).fetchall()
    processed = conn.execute("SELECT * FROM checklists WHERE status='Approved' AND checklist_type=? ORDER BY id DESC LIMIT 200", (ctype,)).fetchall()
    employees = conn.execute("SELECT e_code, employee_name, vertical, sub_vertical, rfds FROM employees ORDER BY employee_name LIMIT 500").fetchall()
    assets = conn.execute("SELECT device_tag, asset_type, serial_number, system_name, username, e_code FROM assets ORDER BY device_tag LIMIT 1000").fetchall()
    software = conn.execute("SELECT desktop_laptop_tag, e_code, software_name, license_key, status FROM software_assignments ORDER BY software_name LIMIT 1000").fetchall()
    conn.close()
    return {"pending": pending, "processed": processed, "employees": employees, "assets": assets, "software": software, "ctype": ctype}


@app.route("/asset-allotment")
@login_required
def asset_allotment():
    return redirect(url_for("asset_allotment_type", ctype="joining"))


@app.route("/asset-allotment/<ctype>", methods=["GET", "POST"])
@login_required
def asset_allotment_type(ctype: str):
    type_map = {"joining": "Joining", "replacement": "Replacement", "rebuild": "Rebuild", "exit": "Exit", "sysadmin": "Sysadmin", "machine-preparation": "Machine Preparation"}
    checklist_type = type_map.get(ctype, "Joining")
    if request.method == "POST" and current_user()["role"] in ("Super Admin", "Admin", "Asset Manager", "Editor"):
        payload = request.form.to_dict(flat=False)
        flat = {k: ", ".join(v) for k, v in payload.items()}
        record = {"checklist_type": checklist_type, "employee_name": request.form.get("employee_name", ""), "e_code": request.form.get("e_code", ""), "vertical": request.form.get("vertical", ""), "sub_vertical": request.form.get("sub_vertical", ""), "rfds": request.form.get("rfds", ""), "payload": json.dumps(flat), "asset_updates": request.form.get("asset_updates", ""), "software_updates": request.form.get("software_updates", ""), "employee_status": request.form.get("employee_status", "Active" if checklist_type != "Exit" else "Exited"), "status": "Pending"}
        rid = insert_record("checklists", CHECKLIST_COLUMNS, record, {"maker": session["username"], "processed": "0"})
        upload_dir = BASE_DIR / "uploads" / "checklists" / str(rid); upload_dir.mkdir(parents=True, exist_ok=True)
        for image in request.files.getlist("images"):
            if image and image.filename:
                target = upload_dir / image.filename; image.save(target)
                execute("INSERT INTO checklist_images (checklist_id,file_path,original_name,created_at) VALUES (?,?,?,?)", (rid, str(target.relative_to(BASE_DIR)), image.filename, now()))
        approver = request.form.get("approver") or "admin"
        if checklist_type in ("Machine Preparation", "Exit"):
            sophos_payload = {"request_type": "Machine Preparation" if checklist_type == "Machine Preparation" else "Exit Process", "hostname": request.form.get("host_name", ""), "new_hostname": request.form.get("host_name", ""), "username": request.form.get("employee_name", ""), "approver": approver, "ticket_id": request.form.get("ticket_id", ""), "payload": json.dumps(flat), "approval_status": "Pending"}
            insert_record("sophos_requests", SOPHOS_REQUEST_COLUMNS, sophos_payload, {"maker": session["username"]})
        create_notification(approver, "Assets Approval", f"{checklist_type} checklist #{rid} needs approval. Changes: {record['asset_updates'] or 'See checklist details'}")
        flash(f"{checklist_type} checklist submitted for approval.", "success")
        return redirect(url_for("asset_allotment_type", ctype=ctype))
    return render_template("asset_allotment.html", **asset_context(checklist_type), section=ctype)


@app.route("/checklist", methods=["GET", "POST"])
@login_required
def checklist():
    if request.method == "POST" and current_user()["role"] in ("Admin", "Editor"):
        ctype = request.form.get("checklist_type", "Joining")
        payload = {k: v for k, v in request.form.items() if k not in CHECKLIST_COLUMNS}
        record = {
            "checklist_type": ctype, "employee_name": request.form.get("employee_name", ""), "e_code": request.form.get("e_code", ""),
            "vertical": request.form.get("vertical", ""), "sub_vertical": request.form.get("sub_vertical", ""), "rfds": request.form.get("rfds", ""),
            "payload": json.dumps(payload), "asset_updates": request.form.get("asset_updates", ""), "software_updates": request.form.get("software_updates", ""),
            "employee_status": request.form.get("employee_status", "Active" if ctype != "Exit" else "Exited"), "status": "Pending"
        }
        rid = insert_record("checklists", CHECKLIST_COLUMNS, record, {"maker": session["username"], "processed": "0"})
        upload_dir = BASE_DIR / "uploads" / "checklists" / str(rid); upload_dir.mkdir(parents=True, exist_ok=True)
        for image in request.files.getlist("images"):
            if image and image.filename:
                target = upload_dir / image.filename
                image.save(target)
                execute("INSERT INTO checklist_images (checklist_id,file_path,original_name,created_at) VALUES (?,?,?,?)", (rid, str(target.relative_to(BASE_DIR)), image.filename, now()))
        create_notification(session["username"], "Assets Approval", f"Checklist #{rid} submitted for approval")
        flash("Checklist submitted for approval.", "success")
        return redirect(url_for("checklist"))
    conn = db()
    pending = conn.execute("SELECT * FROM checklists WHERE status='Pending' ORDER BY id DESC LIMIT 200").fetchall()
    processed = conn.execute("SELECT * FROM checklists WHERE status='Approved' ORDER BY id DESC LIMIT 200").fetchall()
    employees = conn.execute("SELECT e_code, employee_name, vertical, sub_vertical, rfds FROM employees ORDER BY employee_name LIMIT 500").fetchall()
    assets = conn.execute("SELECT * FROM assets ORDER BY device_tag LIMIT 500").fetchall()
    conn.close()
    return render_template("checklist.html", pending=pending, processed=processed, employees=employees, assets=assets)


@app.route("/checklist/upload", methods=["POST"])
@login_required
@roles_required("Admin", "Asset Manager", "Editor")
def checklist_upload():
    ctype = request.form.get("checklist_type", "Joining")
    records = parse_upload(request.files["file"])
    for record in records:
        payload = row_dict(SAMPLE_MAP.get(f"checklist-{ctype.lower()}", []), record)
        insert_record("checklists", CHECKLIST_COLUMNS, {"checklist_type": ctype, "employee_name": payload.get("employee_name", ""), "e_code": payload.get("e_code", ""), "vertical": payload.get("vertical", ""), "sub_vertical": payload.get("sub_vertical", ""), "rfds": payload.get("rfds", ""), "payload": json.dumps(payload), "employee_status": "Active" if ctype != "Exit" else "Exited", "status": "Pending"}, {"maker": session["username"], "processed": "0"})
    flash(f"Uploaded {len(records)} checklist rows.", "success")
    return redirect(url_for("checklist"))


@app.route("/checklist/approve/<int:record_id>", methods=["POST"])
@login_required
@roles_required("Admin", "Asset Manager")
def approve_checklist(record_id: int):
    conn = db(); row = conn.execute("SELECT * FROM checklists WHERE id=?", (record_id,)).fetchone()
    if row:
        if row["e_code"]:
            conn.execute("UPDATE employees SET employee_status=?, updated_at=? WHERE e_code=?", (row["employee_status"] or "Active", now(), row["e_code"]))
            if row["checklist_type"] == "Exit":
                conn.execute("UPDATE assets SET status='Unallocated', e_code='', username='', updated_at=? WHERE e_code=?", (now(), row["e_code"]))
                conn.execute("UPDATE software_assignments SET status='Available', updated_at=? WHERE e_code=?", (now(), row["e_code"]))
        conn.execute("UPDATE checklists SET status='Approved', processed=1, checker=?, updated_at=? WHERE id=?", (session["username"], now(), record_id))
        conn.commit()
    conn.close(); log_action("approve_checklist", "checklists", record_id)
    flash("Checklist approved and inventory references updated.", "success")
    return redirect(url_for("checklist"))


@app.route("/checklist/reject/<int:record_id>", methods=["POST"])
@login_required
@roles_required("Admin", "Asset Manager")
def reject_checklist(record_id: int):
    execute("UPDATE checklists SET status='Rejected', checker=?, updated_at=? WHERE id=?", (session["username"], now(), record_id))
    flash("Checklist rejected.", "warning")
    return redirect(url_for("checklist"))


def calculate_revoke(apply_date: str, duration: str) -> tuple[str, str]:
    try:
        days = int(duration or 0)
        start = datetime.strptime((apply_date or "")[:10], "%Y-%m-%d").date()
        revoke = start + timedelta(days=days)
        return revoke.isoformat(), str((revoke - date.today()).days)
    except (ValueError, TypeError):
        return "", ""


def sync_sophos_machine(conn: sqlite3.Connection, data: dict[str, str]) -> None:
    host = data.get("new_hostname") or data.get("hostname") or data.get("old_hostname")
    if not host:
        return
    existing = conn.execute("SELECT id FROM sophos_machines WHERE hostname=?", (host,)).fetchone()
    payload = {"asset_tag": data.get("asset_tag", ""), "hostname": host, "username": data.get("username") or data.get("requester_name", ""), "ticket_id": data.get("ticket_id", ""), "group_name": data.get("group_name", ""), "application_control": data.get("application_control", ""), "data_loss_prevention": data.get("data_loss_prevention") or data.get("policy", ""), "windows_firewall": data.get("windows_firewall", ""), "peripheral_control": data.get("peripheral_control", ""), "threat_protection": data.get("threat_protection", ""), "update_management": data.get("update_management", ""), "web_control": data.get("web_control", ""), "status": data.get("current_status") or data.get("status", "Active"), "comment": data.get("comment") or data.get("access_comment") or data.get("reason", ""), "updated_at": now()}
    if existing:
        conn.execute("UPDATE sophos_machines SET " + ",".join(f"{k}=?" for k in payload) + " WHERE id=?", tuple(payload.values()) + (existing["id"],))
    else:
        payload["created_at"] = now(); keys = list(payload.keys())
        conn.execute(f"INSERT INTO sophos_machines ({','.join(keys)}) VALUES ({','.join('?' for _ in keys)})", tuple(payload.values()))


@app.route("/sophos", methods=["GET", "POST"])
@login_required
def sophos():
    if request.method == "POST" and current_user()["role"] != "Viewer":
        payload = row_dict(SOPHOS_REQUEST_COLUMNS, request.form)
        if payload.get("request_type") in ("USB Access", "DLP Access", "Remote Access"):
            revoke, days_left = calculate_revoke(payload.get("apply_date", ""), payload.get("duration", ""))
            payload["will_revoke_date"], payload["day_left_for_revoke"] = revoke, days_left
        payload["payload"] = json.dumps(request.form.to_dict())
        payload["approval_status"] = "Pending"
        rid = insert_record("sophos_requests", SOPHOS_REQUEST_COLUMNS, payload, {"maker": session["username"]})
        create_notification(request.form.get("approver") or "admin", "Assets Approval", f"Sophos {payload.get('request_type')} request #{rid} needs approval")
        flash("Sophos request submitted for approval.", "success")
        return redirect(url_for("sophos", section=payload.get("request_type", "machines")))
    section = request.args.get("section", "machines")
    conn = db()
    machines = conn.execute("SELECT * FROM sophos_machines ORDER BY updated_at DESC LIMIT 500").fetchall()
    pending = conn.execute("SELECT * FROM sophos_requests WHERE approval_status='Pending' ORDER BY id DESC LIMIT 300").fetchall()
    requests = conn.execute("SELECT * FROM sophos_requests ORDER BY id DESC LIMIT 500").fetchall()
    section_requests = requests if section in ("machines", "Unprocessed") else [r for r in requests if r["request_type"] == section]
    conn.close()
    return render_template("sophos.html", machines=machines, pending=pending, requests=requests, section_requests=section_requests, columns=SOPHOS_MACHINE_COLUMNS, request_columns=SOPHOS_REQUEST_COLUMNS, section=section)


@app.route("/sophos/machine-action", methods=["POST"])
@login_required
@roles_required("Super Admin", "Admin", "Asset Manager", "Editor")
def sophos_machine_action():
    action = request.form.get("request_type", "USB Access")
    hostname = request.form.get("hostname", "").strip()
    payload = row_dict(SOPHOS_REQUEST_COLUMNS, request.form)
    payload["request_type"] = action
    payload["hostname"] = hostname
    payload["approver"] = request.form.get("approver_1") or request.form.get("approver") or "admin"
    payload["payload"] = json.dumps(request.form.to_dict())
    payload["approval_status"] = "Pending"
    if action in ("USB Access", "DLP Access", "Remote Access"):
        payload["will_revoke_date"], payload["day_left_for_revoke"] = calculate_revoke(payload.get("apply_date", ""), payload.get("duration", ""))
    rid = insert_record("sophos_requests", SOPHOS_REQUEST_COLUMNS, payload, {"maker": session["username"]})
    create_notification(payload["approver"], "Assets Approval", f"Sophos {action} request #{rid} for {hostname} needs approval")
    flash(f"Sophos {action} submitted for approval.", "success")
    return redirect(url_for("sophos", section="Unprocessed"))


@app.route("/sophos/upload", methods=["POST"])
@login_required
@roles_required("Super Admin", "Admin", "Asset Manager", "Editor")
def sophos_upload():
    rtype = request.form.get("request_type", "Machine Preparation")
    records = parse_upload(request.files["file"])
    for record in records:
        payload = row_dict(SOPHOS_REQUEST_COLUMNS, record); payload["request_type"] = rtype; payload["approval_status"] = "Pending"; payload["payload"] = json.dumps(record)
        if rtype in ("USB Access", "DLP Access", "Remote Access"):
            payload["will_revoke_date"], payload["day_left_for_revoke"] = calculate_revoke(payload.get("apply_date", ""), payload.get("duration", ""))
        insert_record("sophos_requests", SOPHOS_REQUEST_COLUMNS, payload, {"maker": session["username"]})
    flash(f"Uploaded {len(records)} Sophos {rtype} requests for approval.", "success")
    return redirect(url_for("sophos", section=rtype))


@app.route("/sophos/approve/<int:record_id>", methods=["POST"])
@login_required
@roles_required("Super Admin", "Admin", "Asset Manager")
def approve_sophos(record_id: int):
    conn = db(); row = conn.execute("SELECT * FROM sophos_requests WHERE id=?", (record_id,)).fetchone()
    if row:
        if row["request_type"] == "Delete":
            conn.execute("DELETE FROM sophos_machines WHERE hostname=? OR hostname=?", (row["hostname"], row["new_hostname"]))
        else:
            sync_sophos_machine(conn, dict(row))
        conn.execute("UPDATE sophos_requests SET approval_status='Approved', checker=?, updated_at=? WHERE id=?", (session["username"], now(), record_id))
        conn.commit()
    conn.close(); log_action("approve_sophos", "sophos_requests", record_id)
    flash("Sophos request approved and machine view updated.", "success")
    return redirect(url_for("sophos", section="Unprocessed"))


@app.route("/sophos/reject/<int:record_id>", methods=["POST"])
@login_required
@roles_required("Super Admin", "Admin", "Asset Manager")
def reject_sophos(record_id: int):
    conn = db(); row = conn.execute("SELECT maker FROM sophos_requests WHERE id=?", (record_id,)).fetchone(); conn.close()
    execute("UPDATE sophos_requests SET approval_status='Rejected', checker=?, updated_at=? WHERE id=?", (session["username"], now(), record_id))
    if row and row["maker"]:
        create_notification(row["maker"], "Change", f"Sophos request #{record_id} was rejected")
    flash("Sophos request rejected and maker notified.", "warning")
    return redirect(url_for("sophos"))


@app.route("/reports/sophos")
@login_required
def sophos_report():
    rtype = request.args.get("type", "")
    conn = db()
    if rtype:
        rows = conn.execute("SELECT * FROM sophos_requests WHERE request_type=? ORDER BY id DESC", (rtype,)).fetchall()
    else:
        rows = conn.execute("SELECT * FROM sophos_requests ORDER BY id DESC").fetchall()
    conn.close()
    if request.args.get("download"):
        return csv_response("Sophos Report", SOPHOS_REQUEST_COLUMNS, rows)
    return render_template("sophos_report.html", rows=rows, columns=SOPHOS_REQUEST_COLUMNS, rtype=rtype)


@app.route("/admin/notifications", methods=["GET", "POST"])
@login_required
@roles_required("Super Admin", "Admin")
def admin_notifications():
    if request.method == "POST":
        execute("DELETE FROM notification_rules WHERE username=?", (request.form["username"],))
        for ntype in request.form.getlist("notification_types"):
            execute("INSERT INTO notification_rules (username,notification_type,enabled,email_enabled,created_at) VALUES (?,?,?,?,?)", (request.form["username"], ntype, 1, 1 if request.form.get(f"email_{ntype}") else 0, now()))
        flash("Notification configuration saved.", "success")
        return redirect(url_for("admin_notifications"))
    conn = db(); users = conn.execute("SELECT username,role,email FROM users ORDER BY username").fetchall(); rules = conn.execute("SELECT * FROM notification_rules ORDER BY username").fetchall(); conn.close()
    return render_template("admin_notifications.html", users=users, rules=rules, notification_types=NOTIFICATION_TYPES)


init_db()

if __name__ == "__main__":
    app.run(debug=True)
